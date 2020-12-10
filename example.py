import discord
from discord.utils import get
from discord.voice_client import VoiceClient
from discord import FFmpegPCMAudio
import asyncio
import random
import openpyxl
import time
import ffmpeg
import youtube_dl
from youtube_dl import YoutubeDL
import crollEx
import pafy

client = discord.Client()
queues = []


@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("----------------")
    await client.change_presence(activity=discord.Game(name="테스트"))


@client.event
async def on_message(message):
    global i, voice
    if message.content.startswith("!"):
        author = message.author
        content = message.content
        print('{}: {}'.format(author, content))

    def pred(m):
        return m.author == message.author and m.channel == message.channel

    if message.content == "!인사":
        author = message.author
        await message.channel.send(author.mention + " " + "안녕하세요")

    if message.content.startswith("!고르기"):
        choice = message.content.split(" ")
        choice_num = random.randint(1, len(choice) - 1)
        choice_res = choice[choice_num]
        await message.channel.send(choice_res)

    if message.content == "!공지":
        file = open("채팅방공지그글.txt")
        await message.channel.send(file.read())
        file.close()

    if message.content.startswith("!사진"):
        i = 1
        img = message.content.split(" ")
        ccl = img[1]
        try:
            ee = int(img[-1])
            for i in range(2, len(img) - 1):
                ccl = ccl + " " + img[i]
                i = i + 1
        except:
            for i in range(2, len(img)):
                ccl = ccl + " " + img[i]
                i = i + 1
        link_img = crollEx.get_image(ccl, img[-1])
        await message.channel.send(link_img)

    if message.content.startswith("!날씨"):
        www = message.content.split(" ")
        author = message.author
        try:
            wea = crollEx.get_location_ko(www[1])
            await message.channel.send(author.mention + " " + wea)
        except:
            await message.channel.send("알 수 없는 지역입니다.")

    if message.content.startswith("!학습"):
        file = openpyxl.load_workbook("디스코드.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 501):
            if sheet["A" + str(i)].value == learn[1] and sheet["B" + str(i)].value == learn[2]:
                await message.channel.send("이미 학습되어 있는 단어입니다.")
                break
            elif sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("학습되었습니다.")
                break
        file.save("디스코드.xlsx")

    if message.content.startswith("!덮어쓰기"):
        file = openpyxl.load_workbook("디스코드.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 501):
            if sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("학습되었습니다.")
                break
            elif sheet["A" + str(i)].value != learn[1]:
                await message.channel.send("해당 단어는 학습된 적이 없거나 잊었습니다.")
                break
            else:
                await message.channel.send("입력 오류입니다.")
                break
        file.save("디스코드.xlsx")

    if message.content.startswith("!기억") and not message.content.startswith("!기억삭제"):
        file = openpyxl.load_workbook("디스코드.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 501):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break
        if sheet["A" + str(i)].value != memory[1]:
            await message.channel.send("단어를 학습한 적이 없거나 잊었습니다.")

    if message.content.startswith("!기억삭제"):
        file = openpyxl.load_workbook("디스코드.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 501):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                await message.channel.send("단어를 잊었습니다.")
                file.save("디스코드.xlsx")
                break

    if message.content.startswith("!대리할말"):
        con = message.content.split(" ")
        tok = int(con[1])
        author = message.guild.get_member(tok)
        msg = con[2]
        await message.channel.send(author.mention + msg)

    if message.content.startswith("!초대"):
        await message.author.send("https://discordapp.com/oauth2/authorize?client_id=531341952248053775&scope=bot")

    if message.content.startswith("!대화시작"):
        await message.author.send("반갑습니다.")

    if message.content.startswith("!대리귓말"):
        con = message.content.split(" ")
        tok = int(con[1])
        author = message.guild.get_member(tok)
        msg = con[2]
        await author.send(msg)

    if message.content == "!참가":
        channel = message.author.voice.channel
        if channel is not None:
            try:
                await channel.connect()
            except discord.ClientException:
                await message.channel.send("이미 음성채널에 연결되었습니다.")
        else:
            await message.channel.send("호출 대상이 음청채널에 없습니다.")

    if message.content == "!떠나기":
        guild = message.guild
        try:
            voice_client = guild.voice_client
            await voice_client.disconnect()
        except:
            await message.channel.send("이미 음성채널에서 연결이 해제되었습니다.")

    if message.content.startswith("!재생"):
        sentence = message.content.split(" ")
        url = sentence[1]
        guild = message.guild
        FFMPEG_OPTS = {'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5', 'options': '-vn'}
        video, source = await search(url)
        voice = discord.utils.get(client.voice_clients, guild=guild)
        voice.play(FFmpegPCMAudio(source, **FFMPEG_OPTS), after=lambda e: print('done', e))
        voice.is_playing()

    if message.content == "!일시정지":
        voice.pause()

    if message.content == "!정지":
        voice.stop()

    if message.content == "!재개":
        voice.resume()

    if message.content.startswith("!볼륨"):
        sentence = message.content.split(" ")
        vol = sentence[1]
        guild = message.guild
        voice_client = guild.voice_client
        voice_client.source = discord.PCMVolumeTransformer(voice_client.source)
        voice_client.source.volume = vol

    if message.content == "!건너뛰기":
        voice.stop()
        await check_queue()

    if message.content.startswith("!예약") and not message.content.startswith("!예약목록"):
        sentence = message.content.split(" ")
        url = sentence[1]
        queues.append(url)
        await message.channel.send("예약되었습니다.")

    if message.content == "!예약목록":
        await message.channel.send("예약 개수는 " + str(len(queues)) + "개 입니다.")

    if message.content == "!리스트확인":
        queues.append("a")
        await message.channel.send(queues)
        queues.pop()

    if message.content == "!종료":
        exit(0)

    if message.content == "!집사 꺼져":
        await message.channel.send("ㅠㅠ")
        exit(0)

    if message.content.startswith("!@CM"):
        ch = message.content.split(" ")[1]
        msg = message.content.split(" ")[2]
        for i in range(3, len(message.content.split(" "))):
            msg = msg + " " + message.content.split(" ")[i]
            i = i + 1
        await client.get_channel(int(ch)).send(msg)


@client.event
async def check_queue(message):
    if queues:
        print("ok")

@client.event
async def search(query):
    with YoutubeDL({'format': 'bestaudio', 'noplaylist':'True'}) as ydl:
        info = ydl.extract_info(query, download=False)
    return (info, info['formats'][0]['url'])

client.run("NTMxMzQxOTUyMjQ4MDUzNzc1.DxNXaA.E1p7SLcaPE2sDTnbruRwvdnj0b4")
